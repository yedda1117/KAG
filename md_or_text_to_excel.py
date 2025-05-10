# Requires:
# pip install openai openpyxl

# 我已将脚本改为支持多行命令行输入，并用 argparse 提供两种模式：

# --input-file / -i：读取指定 Markdown 文件

# 无 -i 时：从标准输入中读入多行（结束输入请按 Ctrl+D）
# 无 -k 时：调用设定好的API_KEY

# 示例用法：

# 从文件读取
# python script.py -i question.md

# 交互粘贴多行
# python script.py
# # 程序会提示“Paste your markdown...”，粘贴完按 Ctrl+D 即可
# Excel 路径可通过 --excel-path（或 -o）自定义，默认为 questions_log.xlsx。


import json
import sys
import argparse
from openai import OpenAI
from json import JSONDecodeError
from openpyxl import Workbook, load_workbook
import os


# Initialize client
def init_client(api_key: str, base_url: str = "https://api.deepseek.com") -> OpenAI:
    return OpenAI(api_key=api_key, base_url=base_url)


# Clean code fences from LLM output
def clean_fences(raw: str) -> str:
    lines = raw.strip().splitlines()
    # Remove starting ```json or ```
    if lines and lines[0].startswith("```"):
        lines = lines[1:]
    # Remove ending ```
    if lines and lines[-1].startswith("```"):
        lines = lines[:-1]
    return "\n".join(lines)


# Call LLM and enforce JSON-only response
def extract_structured(raw_text: str, client: OpenAI) -> dict:
    # 首先我们需要修改这个，这个是系统级别的提示词（优先级最高，对于该级别的提示词，模型相应最准确）
    # 为什么这里需要用英文，因为英文似乎效果会更好（实验得出，无理论支撑）
    system_msg = (
        "You are a JSON-only assistant.\n"
        "Extract and clean the question, answer and answer_analysis, determine knowledge_type, description_knowledge, question_type, teaching_hints and difficulty.\n"
        "Knowledge_type according to the hierarchical format, the language needs to be Chinese: Major class :: Middle class :: minor class (such as 数学 :: 数列 :: 等差数列)\n"
        "Please briefly describe knowledge according to knowledge_type into description_knowledge involved in the question."
        "Question, answer and answer_analysis, the text part needs to be in Chinese except for the necessary English parts, and the mathematical symbols and formulas part uses latex syntax.\n"
        "Difficulty ratings are integers from 1 to 5 (1= easiest, 5= hardest)\n"
        "In question_type the output language should be Chinese (such as 选择题, 填空题, 解答题, etc.)\n"
        "In teaching_hints should include core problem-solving strategies in Chinese.\n"
        "Always output valid JSON with keys: difficulty, knowledge_type, question, answer_analysis, answer, question_type, description_knowledge, teaching_hints.\n"
        "Do not include any additional text.\n"
        "If a field cannot be extracted, set it to an empty string."
        "Please strictly follow the following JSON format to ensure that: 1. Use double quotes for all strings. 2. All special characters are escaped correctly. 3. Not using single quotes or unescaped backslashes."
    )
    user_msg = f"Process this text:\n```{raw_text}```"

    # Stream the response for real-time display
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        stream=True,
    )

    full_output = ""
    print("Streaming LLM output:")
    for chunk in response:
        delta = None
        # Handle different OpenAI SDK structures
        try:
            delta = chunk.choices[0].delta.content
        except AttributeError:
            delta = chunk.choices[0].message.content
        if delta:
            print(delta, end="", flush=True)
            full_output += delta
    print()  # newline after stream

    # Clean fences and parse JSON
    json_str = clean_fences(full_output)
    try:
        data = json.loads(json_str)
    except JSONDecodeError as e:
        raise ValueError(
            f"Invalid JSON output from LLM: {e}\nFull output was: {full_output}"
        )
    # Ensure all expected keys
    # difficulty, knowledge_type, question,
    # answer_analysis, answer, question_type, teaching_hints.
    for key in [
        "difficulty",
        "knowledge_type",
        "question",
        "answer_analysis",
        "answer",
        "teaching_hints",
        "question_type",
        "description_knowledge",
        "teaching_hints",
    ]:
        if key not in data:
            raise KeyError(f"Missing key in JSON output: {key}")
    return data


# Update or create Excel file
def update_excel(json_data: dict, filepath: str) -> None:
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # write header
        ws.append(list(json_data.keys()))
    # append row of values in same order as header
    header = [cell.value for cell in ws[1]]
    row = [json_data.get(col, "") for col in header]
    ws.append(row)
    wb.save(filepath)


# Main entry point
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extract structured Q&A from markdown and log to Excel."
    )
    parser.add_argument(
        "--api-key",
        "-k",
        default="sk-9939d8ab6f0441a3a255c057c1ba0552",
        help="DeepSeek API key",
    )
    parser.add_argument(
        "--base-url",
        "-b",
        default="https://api.deepseek.com/",
        help="DeepSeek API base URL",
    )
    parser.add_argument(
        "--input-file", "-i", help="Path to markdown file; if omitted, reads from stdin"
    )
    parser.add_argument(
        "--excel-path",
        "-o",
        default="questions_log.xlsx",
        help="Path to output Excel file",
    )
    args = parser.parse_args()

    # Read raw_text either from file or stdin for multiline input
    # if args.input_file:
    #     with open(args.input_file, "r", encoding="utf-8") as f:
    #         raw_text = f.read()
    # else:
    #     print("Paste your markdown text and press Ctrl+Z (EOF) when done:")
    #     raw_text = sys.stdin.read()

    client = init_client(args.api_key, args.base_url)

    while True:
        print("Paste your markdown text and press Ctrl+Z (EOF) when done:")
        raw_text = sys.stdin.read()
        try:
            structured = extract_structured(raw_text, client)
            print(
                "Extracted JSON:\n",
                json.dumps(structured, ensure_ascii=False, indent=2),
            )
        except Exception as e:
            print("Error during JSON extraction:", e)
            sys.exit(1)

        update_excel(structured, args.excel_path)
        print(f"Updated Excel file: {args.excel_path}")
